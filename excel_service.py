# -*- coding: utf-8 -*-
"""
excel_service.py
"""

import os
import re

import pandas as pd
from openpyxl import load_workbook

from constants import HEADER_SCAN_ROWS, LISTING_SCAN_ROWS
from helpers import (
    float_to_br,
    format_product_code,
    maybe_number_to_br,
    money_to_float,
    normalize_text,
    safe_str,
)


def detect_engine(path):
    ext = os.path.splitext(path)[1].lower()
    if ext == ".xlsx":
        return "openpyxl"
    if ext == ".xls":
        return "xlrd"
    return None


def list_excel_files(folder):
    out = []
    if not os.path.isdir(folder):
        return out
    for name in os.listdir(folder):
        if name.startswith("~$"):
            continue
        low = name.lower()
        if low.endswith(".xlsx") or low.endswith(".xls"):
            out.append(os.path.join(folder, name))
    out.sort()
    return out


def extract_store_number(text):
    s = safe_str(text)

    m = re.search(r"\bLOJA\s*0*([0-9]+)\b", s, flags=re.I)
    if m:
        return m.group(1)

    m = re.search(r"\bFILIAL\s*0*([0-9]+)\b", s, flags=re.I)
    if m:
        return m.group(1)

    m = re.search(r"\b0*([0-9]{1,4})\b", s)
    if m:
        return str(int(m.group(1)))

    return s


def find_best_sheet(path):
    engine = detect_engine(path)
    xls = pd.ExcelFile(path, engine=engine)
    preferred = ["Relatorio", "RELATORIO", "Relatório", "RELATÓRIO"]
    for s in xls.sheet_names:
        if s in preferred:
            return s
    for s in xls.sheet_names:
        if "RELAT" in s.upper():
            return s
    return xls.sheet_names[0]


def row_score_for_header(values):
    vals = [normalize_text(v) for v in values]
    joined = " | ".join([v for v in vals if v])

    score = 0
    tokens = [
        "CODIGO", "CÓDIGO", "DESCRICAO", "DESCRIÇÃO", "UM",
        "AT. VENDA", "AT VENDA",
        "ULT. COMPRA", "ULT COMPRA",
        "/ COMPRA", "COMPRA",
        "CUSTO",
        "ULT. VENDA", "ULT VENDA",
        "/ VENDA", "VENDA",
        "SUGESTAO", "SUGESTÃO",
        "MARG", "MARG. PAD", "MARG PAD",
        "DT. FUTURA", "DT FUTURA", "FUTURA"
    ]
    for t in tokens:
        if t in joined:
            score += 1

    if "CODIGO" in joined and ("DESCRICAO" in joined or "DESCRIÇÃO" in joined):
        score += 4
    if "SUGESTAO" in joined or "SUGESTÃO" in joined:
        score += 2
    if "MARG" in joined:
        score += 2

    return score


def detect_header_row(raw_df):
    best_idx = None
    best_score = -1

    limit = min(HEADER_SCAN_ROWS, len(raw_df))
    for i in range(limit):
        score = row_score_for_header(raw_df.iloc[i].tolist())
        if score > best_score:
            best_score = score
            best_idx = i

    if best_idx is None or best_score < 4:
        raise ValueError("Não foi possível localizar a linha de cabeçalho da planilha.")

    return best_idx


def normalize_detected_headers(header_values):
    headers = [safe_str(v) for v in header_values]

    fixed = []
    for idx, h in enumerate(headers):
        up = normalize_text(h)

        if up in ("CODIGO", "CÓDIGO"):
            fixed.append("CODIGO")
        elif up in ("DESCRICAO", "DESCRIÇÃO"):
            fixed.append("DESCRICAO")
        elif up == "UM":
            fixed.append("UM")
        elif "AT. VENDA" in up or "AT VENDA" in up:
            fixed.append("AT_VENDA")
        elif "ULT. COMPRA" in up or "ULT COMPRA" in up:
            fixed.append("ULT_COMPRA")
        elif up == "/ COMPRA":
            fixed.append("COMPRA_SEP")
        elif up == "COMPRA" and idx > 0:
            fixed.append("COMPRA_SEP")
        elif up == "CUSTO":
            fixed.append("CUSTO")
        elif "ULT. VENDA" in up or "ULT VENDA" in up:
            fixed.append("ULT_VENDA")
        elif up == "/ VENDA":
            fixed.append("VENDA_SEP")
        elif up == "VENDA" and idx > 0:
            fixed.append("VENDA_SEP")
        elif "SUGESTAO" in up or "SUGESTÃO" in up:
            fixed.append("SUGESTAO")
        elif "MARG. PAD" in up or "MARG PAD" in up:
            fixed.append("MARG_PAD")
        elif "MARG %" in up or up == "MARG" or up == "MARGEM" or "% MARG" in up:
            fixed.append("MARGEM")
        elif "DT. FUTURA" in up or "DT FUTURA" in up or "FUTURA" in up:
            fixed.append("DT_FUTURA")
        elif "NOVO_PRECO_EDITADO" in up or "NOVO PRECO EDITADO" in up or "NOVO_PRECO" in up or "NOVO PRECO" in up:
            fixed.append("NOVO_PRECO_EDITADO")
        else:
            fixed.append(h if h else f"COL_{idx}")

    counts = {}
    unique = []
    for h in fixed:
        base = h or "COL"
        counts[base] = counts.get(base, 0) + 1
        if counts[base] == 1:
            unique.append(base)
        else:
            unique.append(f"{base}_{counts[base]}")

    return unique


def build_column_map(df):
    mapping = {}
    for c in df.columns:
        mapping[normalize_text(c)] = c
    return mapping


def pick_col(colmap, *names):
    for n in names:
        if n in colmap:
            return colmap[n]
    return None


def detect_store_name_from_raw_df(raw_df, filepath):
    limit = min(LISTING_SCAN_ROWS, len(raw_df))
    for i in range(limit):
        row_vals = [safe_str(x) for x in raw_df.iloc[i].tolist() if safe_str(x)]
        joined = " ".join(row_vals)

        m = re.search(r"Empresa/Filial:\s*([0-9]+)", joined, flags=re.I)
        if m:
            return m.group(1).strip()

        m2 = re.search(r"\bLOJA\s*0*([0-9]+)\b", joined, flags=re.I)
        if m2:
            return m2.group(1).strip()

    if len(raw_df) > 2:
        c0 = safe_str(raw_df.iloc[2, 0])
        if c0:
            return extract_store_number(c0)

    base = os.path.splitext(os.path.basename(filepath))[0]
    base = re.sub(r"[_\-]+", " ", base)
    return extract_store_number(base.strip().upper())


def scan_store_metadata(path):
    engine = detect_engine(path)
    if not engine:
        raise ValueError(f"Formato não suportado: {path}")

    sheet = find_best_sheet(path)
    raw_df = pd.read_excel(path, sheet_name=sheet, engine=engine, header=None, nrows=LISTING_SCAN_ROWS)
    raw_df = raw_df.dropna(how="all")
    raw_df = raw_df.loc[:, ~raw_df.columns.astype(str).str.contains("^Unnamed", case=False, na=False)]
    raw_df = raw_df.copy()

    store_name = detect_store_name_from_raw_df(raw_df, path)
    store_num = extract_store_number(store_name)

    return {
        "arquivo": path,
        "sheet": sheet,
        "loja": store_num,
        "engine": engine,
        "dirty": False,
        "loaded": False,
        "load_error": None,
        "header_row": None,
        "rows_count": 0,
        "changed_count": 0,
        "visited": False,
        "hidden": False,
    }


def load_full_store_data(meta, edits_map):
    path = meta["arquivo"]
    sheet = meta["sheet"]
    engine = meta["engine"]

    raw_df = pd.read_excel(path, sheet_name=sheet, engine=engine, header=None)
    raw_df = raw_df.dropna(how="all")
    raw_df = raw_df.loc[:, ~raw_df.columns.astype(str).str.contains("^Unnamed", case=False, na=False)]
    raw_df = raw_df.copy()

    header_row = detect_header_row(raw_df)
    headers = normalize_detected_headers(raw_df.iloc[header_row].tolist())

    df = raw_df.iloc[header_row + 1:].copy()
    df.columns = headers
    df = df.dropna(how="all")
    df = df.reset_index(drop=False).rename(columns={"index": "__source_index__"})

    colmap = build_column_map(df)

    code_col = pick_col(colmap, "CODIGO", "CÓDIGO", "COD", "ID")
    desc_col = pick_col(colmap, "DESCRICAO", "DESCRIÇÃO", "PRODUTO", "ITEM", "NOME")
    um_col = pick_col(colmap, "UM")
    at_venda_col = pick_col(colmap, "AT_VENDA")
    ult_compra_col = pick_col(colmap, "ULT_COMPRA")
    compra_sep_col = pick_col(colmap, "COMPRA_SEP")
    custo_col = pick_col(colmap, "CUSTO")
    ult_venda_col = pick_col(colmap, "ULT_VENDA")
    venda_sep_col = pick_col(colmap, "VENDA_SEP")
    preco_sugerido_col = pick_col(colmap, "SUGESTAO", "PRECO_SUGERIDO", "PREÇO_SUGERIDO")
    margem_col = pick_col(colmap, "MARGEM", "MARG")
    margem_padrao_col = pick_col(colmap, "MARG_PAD", "%MARG. PAD.", "%MARG PAD.")
    dt_futura_col = pick_col(colmap, "DT_FUTURA")
    novo_preco_col = pick_col(colmap, "NOVO_PRECO_EDITADO", "NOVO PRECO EDITADO", "NOVO_PRECO", "NOVO PRECO")

    if not code_col and not desc_col:
        raise ValueError(
            "Não foi possível identificar as colunas principais.\n"
            f"Cabeçalhos detectados: {list(df.columns)}"
        )

    rows = []
    rows_map = {}
    changed_count = 0
    file_edits = edits_map.get(path, {})

    for _, row in df.iterrows():
        codigo = format_product_code(row[code_col]) if code_col else ""
        descricao = safe_str(row[desc_col]) if desc_col else ""

        if not codigo and not descricao:
            continue

        dt_futura = row[dt_futura_col] if dt_futura_col else ""
        if hasattr(dt_futura, "strftime"):
            dt_futura = dt_futura.strftime("%d/%m/%Y")
        else:
            dt_futura = safe_str(dt_futura)

        src_idx = int(row["__source_index__"])

        novo_original = money_to_float(row[novo_preco_col]) if novo_preco_col else None
        novo_memoria = file_edits.get(src_idx, novo_original)

        item = {
            "arquivo_origem": path,
            "sheet_origem": sheet,
            "loja": meta["loja"],
            "codigo": codigo,
            "descricao": descricao,
            "um": safe_str(row[um_col]) if um_col else "",
            "at_venda": money_to_float(row[at_venda_col]) if at_venda_col else None,
            "ult_compra": money_to_float(row[ult_compra_col]) if ult_compra_col else None,
            "compra_sep": maybe_number_to_br(row[compra_sep_col]) if compra_sep_col else "",
            "custo": money_to_float(row[custo_col]) if custo_col else None,
            "ult_venda": money_to_float(row[ult_venda_col]) if ult_venda_col else None,
            "venda_sep": maybe_number_to_br(row[venda_sep_col]) if venda_sep_col else "",
            "preco_sugerido": money_to_float(row[preco_sugerido_col]) if preco_sugerido_col else None,
            "margem": money_to_float(row[margem_col]) if margem_col else None,
            "margem_padrao": money_to_float(row[margem_padrao_col]) if margem_padrao_col else None,
            "dt_futura": dt_futura,
            "novo_preco_editado": novo_memoria,
            "alterado": bool(novo_memoria is not None),
            "source_index": src_idx,
        }

        if novo_memoria is not None:
            changed_count += 1

        rows.append(item)
        rows_map[src_idx] = item

    full_data = {
        "arquivo": path,
        "sheet": sheet,
        "loja": meta["loja"],
        "engine": engine,
        "dirty": meta.get("dirty", False),
        "header_row": header_row,
        "rows": rows,
        "rows_map": rows_map,
        "rows_count": len(rows),
        "changed_count": changed_count,
    }
    return full_data


def _save_xlsx_direct(meta, edits_for_file):
    path = meta["arquivo"]
    sheet_name = meta["sheet"]

    wb = load_workbook(path)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Aba não encontrada: {sheet_name}")

    ws = wb[sheet_name]

    header_row_zero = meta.get("header_row")
    if header_row_zero is None:
        raw_df = pd.read_excel(path, sheet_name=sheet_name, engine="openpyxl", header=None)
        header_row_zero = detect_header_row(raw_df)

    header_excel_row = header_row_zero + 1

    novo_col = None
    max_col = ws.max_column

    for c in range(1, max_col + 1):
        val = normalize_text(ws.cell(row=header_excel_row, column=c).value)
        if val == "NOVO_PRECO_EDITADO":
            novo_col = c
            break

    if novo_col is None:
        if max_col >= 16384:
            raise ValueError(
                "A planilha já está no limite máximo de colunas do Excel e não tem a coluna NOVO_PRECO_EDITADO."
            )
        novo_col = max_col + 1
        ws.cell(row=header_excel_row, column=novo_col, value="NOVO_PRECO_EDITADO")

    for src_idx, novo_valor in edits_for_file.items():
        excel_row = int(src_idx) + 1
        ws.cell(
            row=excel_row,
            column=novo_col,
            value=float_to_br(novo_valor) if novo_valor is not None else ""
        )

    wb.save(path)
    wb.close()
    return path, "original"


def persist_store_changes(meta, edits_for_file):
    path = meta["arquivo"]
    ext = os.path.splitext(path)[1].lower()

    if ext == ".xlsx":
        return _save_xlsx_direct(meta, edits_for_file)

    engine = detect_engine(path)
    sheet = meta["sheet"]

    raw_df = pd.read_excel(path, sheet_name=sheet, engine=engine, header=None, dtype=object)
    raw_df = raw_df.copy()

    header_row = detect_header_row(raw_df)

    if raw_df.shape[1] >= 16384:
        header_values = raw_df.iloc[header_row].tolist()
        raw_headers_norm = [normalize_text(h) for h in header_values]
        if "NOVO_PRECO_EDITADO" not in raw_headers_norm:
            raise ValueError(
                "A planilha .xls já está no limite máximo de colunas e não tem a coluna NOVO_PRECO_EDITADO."
            )

    data_start = header_row + 1
    header_values = raw_df.iloc[header_row].tolist()
    headers = normalize_detected_headers(header_values)

    body_df = raw_df.iloc[data_start:].copy()
    body_df.columns = headers
    body_df = body_df.reset_index(drop=False).rename(columns={"index": "__abs_row__"})

    colmap = build_column_map(body_df)
    novo_preco_col = pick_col(colmap, "NOVO_PRECO_EDITADO", "NOVO PRECO EDITADO", "NOVO_PRECO", "NOVO PRECO")

    if not novo_preco_col:
        novo_preco_col = "NOVO_PRECO_EDITADO"
        body_df[novo_preco_col] = ""

    for idx, novo_valor in edits_for_file.items():
        match = body_df["__abs_row__"] == idx
        if match.any():
            body_df.loc[match, novo_preco_col] = float_to_br(novo_valor) if novo_valor is not None else ""

    raw_headers_norm = [normalize_text(h) for h in header_values]

    if "NOVO_PRECO_EDITADO" in raw_headers_norm:
        novo_preco_raw_col = raw_headers_norm.index("NOVO_PRECO_EDITADO")
    else:
        novo_preco_raw_col = raw_df.shape[1]
        raw_df[novo_preco_raw_col] = None
        raw_df.iloc[header_row, novo_preco_raw_col] = "NOVO_PRECO_EDITADO"

    for _, row in body_df.iterrows():
        abs_row = int(row["__abs_row__"])
        val = safe_str(row.get(novo_preco_col, ""))
        raw_df.iat[abs_row, novo_preco_raw_col] = val

    out_path = os.path.splitext(path)[0] + "_EDITADO.xlsx"
    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        raw_df.to_excel(writer, sheet_name=sheet, index=False, header=False)
    return out_path, "copy"